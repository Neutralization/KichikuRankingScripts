# -*- coding: utf-8 -*-
import asyncio
import json
from functools import reduce
from os import listdir

import aiohttp
import arrow
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PAST = arrow.get("2021-11-08 00:00:00", "YYYY-MM-DD HH:mm:ss")
NOW = arrow.now()
weekday = int(NOW.format("d"))
start_date = NOW.shift(days=-(weekday + 6)).format("YYYY-MM-DD")
end_date = NOW.shift(days=-(weekday - 1)).format("YYYY-MM-DD")
weeks = int((NOW.timestamp() - PAST.timestamp()) / 3600 / 24 / 7) + 1

# https://bugs.python.org/issue29288
"".encode("idna")


def find(week, name):
    for files in listdir("."):
        if files.endswith(".xlsx") and week in files and name in files:
            print(f"\t找到{week}期{name}Excel文件")
            return files
    print(f"\t未找到{week}期{name}Excel文件")
    return None


def readExcel(filename):
    wb = load_workbook(filename, data_only=True)
    ws = wb.active
    xlsx_data = []
    last_col = len(list(ws.columns))
    last_row = len(list(ws.rows))
    for row in range(1, last_row + 1):
        data_cols = {}
        for column in range(1, last_col + 1):
            col_num = get_column_letter(column)
            if row > 1:
                data_cols[ws[col_num + str(1)].value] = ws[col_num + str(row)].value
        xlsx_data.append(data_cols)
    xlsx_data = list(filter(lambda x: x != {} and set(x.values()) != {None}, xlsx_data))
    # print([x["bvid"] for x in xlsx_data if x["bvid"] != "bvid"])
    return [x["bvid"] for x in xlsx_data if x["bvid"] != "bvid"]


def writeExcel(filename, data):
    wb = load_workbook(filename)
    ws = wb.active
    last_col = len(list(ws.columns))
    last_row = len(list(ws.rows))
    bvid_col = "A"
    title_col = "A"
    for column in range(1, last_col + 1):
        col_num = get_column_letter(column)
        if ws[col_num + str(1)].value == "bvid":
            bvid_col = col_num
        if ws[col_num + str(1)].value == "title":
            title_col = col_num
    for row in range(1, last_row + 1):
        if ws[bvid_col + str(row)].value != "bvid":
            if data[ws[bvid_col + str(row)].value] != "":
                ws[title_col + str(row)] = data[ws[bvid_col + str(row)].value]
            else:
                continue
    wb.save(filename)


async def getVideoTitle(bvid):
    params = (("bvid", bvid),)
    async with aiohttp.ClientSession() as session:
        async with session.get(
            "https://api.bilibili.com/x/web-interface/view", params=params
        ) as resp:
            content = await resp.text()
            result = json.loads(content)
    if result.get("code") == 0:
        title = result["data"]["title"]
        return {bvid: title}
    else:
        return {bvid: ""}


def main():
    main_excel = find(f"{weeks:03d}", "主榜")
    if main_excel is not None:
        ranks = readExcel(main_excel)
        tasks = [asyncio.ensure_future(getVideoTitle(bvid)) for bvid in ranks]
        loop = asyncio.get_event_loop()
        VideoTitles = loop.run_until_complete(asyncio.gather(*tasks))
        VideoTitleDict = reduce(lambda x, y: {**x, **y}, VideoTitles)
        # print(VideoTitleDict)
        writeExcel(main_excel, VideoTitleDict)
    input("\n\t现在可以关闭本程序\n")


if __name__ == "__main__":
    main()

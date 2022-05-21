# -*- coding: utf-8 -*-
from json import dumps
from os.path import abspath
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def excel2json(filename, ranktype):
    wb = load_workbook(filename, data_only=True)
    ws = wb.active
    xlsx_data = []
    last_col = len(list(ws.columns))
    last_row = len(list(ws.rows))
    for row in range(1, last_row + 1):
        data_cols = {}
        for column in range(1, last_col + 1):
            col_num = get_column_letter(column)
            if ws[col_num + str(1)].value == ws[col_num + str(row)].value:
                break
            if row > 1:
                data_cols[ws[col_num + str(1)].value] = ws[col_num + str(row)].value
        if len(data_cols) != 0:
            xlsx_data.append(data_cols)
    # print(list(map(lambda x: x.get("排名"), xlsx_data[:20])))
    xlsx_data = list(filter(lambda x: x != {} and set(x.values()) != {None}, xlsx_data))
    with open("./psdownload/download.txt", "a", encoding="utf-8") as f:
        if ranktype == "主榜":
            videos = [x["aid"] for x in xlsx_data[:20] if isinstance(x["排名"], int)]
            list(map(lambda x: f.write(f"av{x}\n"), videos))
        elif ranktype == "榜外":
            videos = [x["AV号"] for x in xlsx_data]
            list(map(lambda x: f.write(f"{x.lower()}\n"), videos))
        elif ranktype == "新人":
            videos = [x["AV号"] for x in xlsx_data]
            list(map(lambda x: f.write(f"{x.lower()}\n"), videos))
    if ranktype == "主榜":
        point_data = {
            x["排名"]: x["总分"]
            for x in xlsx_data[:21]
            if isinstance(x["排名"], int)  # and x["排名"] <= 21
        }
        json_data = [
            {
                "rank": n + 1,
                "type": "主榜",
                "video": f"./主榜视频/av{x['aid']}.mp4",
                "image": f"./主榜3-1/Rank_{n+1}.png"
                if n + 1 <= 3
                else f"./主榜10-4/Rank_{n+1-3}.png"
                if 3 < n + 1 <= 10
                else f"./主榜20-11/Rank_{n+1-10}.png",
                "point": format(
                    int(point_data[x["排名"]]) - int(point_data[x["排名"] + 1]), ","
                )
                if point_data.get(x["排名"] + 1) and n + 1 <= 3
                else "000,000,000"
                if not point_data.get(x["排名"] + 1) and n + 1 <= 3
                else "",
                "offset": 0,
            }
            for n, x in enumerate(xlsx_data[:20])
        ]
    elif ranktype == "榜外":
        json_data = [
            {
                "rank": 10 - n,
                "type": "榜外",
                "video": f"./主榜视频/{x['AV号'].lower()}.mp4",
                "image": f"./鬼畜剧场月刊/{n + 1}.png",
                "point": "",
                "offset": 0,
            }
            for n, x in enumerate(xlsx_data)
        ]
    elif ranktype == "新人":
        json_data = [
            {
                "rank": n + 1,
                "type": "新人",
                "video": f"./主榜视频/{x['AV号'].lower()}.mp4",
                "image": f"./鬼畜月刊新人自荐/{n + 1}.png",
                "point": "",
                "offset": 0,
            }
            for n, x in enumerate(xlsx_data)
        ]
    else:
        json_data = []
    return json_data


def main():
    MainExcel = input("\n\t拖拽主榜Excel到窗口，出现地址后回车\n\t")
    print(f"\t{MainExcel}")
    SubExcel = input("\n\t拖拽榜外推荐Excel到窗口，出现地址后回车\n\t")
    print(f"\t{SubExcel}")
    NewExcel = input("\n\t拖拽新人自荐Excel到窗口，出现地址后回车\n\t")
    print(f"\t{NewExcel}")

    json_data = []
    json_data += excel2json(MainExcel, "主榜")
    json_data += excel2json(SubExcel, "榜外")
    json_data += excel2json(NewExcel, "新人")
    print(f"\n\t视频下载列表已保存至“{abspath('./psdownload/download.txt')}”")
    data = dumps(json_data, indent=4, ensure_ascii=False)
    with open("月刊数据.json", "w", encoding="utf-8") as f:
        f.write(data)
    print("\n\tAE脚本数据“月刊数据.json”已经生成")
    input("\n\t可以正常退出...\n\t")


if __name__ == "__main__":
    main()

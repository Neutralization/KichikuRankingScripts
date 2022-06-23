# -*- coding: utf-8 -*-
from json import dumps
from os import listdir
from os.path import abspath

import arrow
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PAST = arrow.get("2021-11-08 00:00:00", "YYYY-MM-DD HH:mm:ss")
NOW = arrow.now()
weekday = int(NOW.format("d"))
start_date = NOW.shift(days=-(weekday + 6)).format("YYYY-MM-DD")
end_date = NOW.shift(days=-(weekday - 1)).format("YYYY-MM-DD")
weeks = int((NOW.timestamp() - PAST.timestamp()) / 3600 / 24 / 7) + 1


def find(week, name):
    for files in listdir("."):
        if files.endswith(".xlsx") and week in files and name in files:
            print(f"\t找到{week}期{name}Excel文件")
            return files
    print(f"\t未找到{week}期{name}Excel文件")
    return None


def an2cn(num):
    a2c = ["零", "一", "二", "三", "四", "五", "六", "七", "八", "九"]
    if num // 10 < 1:
        return a2c[num % 10]
    if num // 10 == 1:
        return "十" + a2c[num % 10] if num % 10 != 0 else "十"
    else:
        return (
            a2c[num // 10] + "十" + a2c[num % 10]
            if num % 10 != 0
            else a2c[num // 10] + "十"
        )


def xlsx2json(filename, ranktype):
    wb = load_workbook(filename, data_only=True)
    ws = wb.active
    xlsx_data = []
    last_col = len(list(ws.columns))
    last_row = len(list(ws.rows))
    for row in range(1, last_row + 1):
        data_cols = {}
        for column in range(1, last_col + 1):
            col_num = get_column_letter(column)
            if ws[col_num + str(1)].value == ws[col_num + str(row)].value:
                break
            if row > 1:
                data_cols[ws[col_num + str(1)].value] = ws[col_num + str(row)].value
        if len(data_cols) != 0:
            xlsx_data.append(data_cols)
    # print(list(map(lambda x: x.get("排名"), xlsx_data[:20])))
    xlsx_data = list(filter(lambda x: x != {} and set(x.values()) != {None}, xlsx_data))

    with open("./psdownload/download.txt", "a", encoding="utf-8") as f:
        if ranktype == "主榜":
            videos = [
                x["aid"]
                for x in xlsx_data[:20]
                if isinstance(x["排名"], int)  # and x["排名"] <= 20
            ]
            list(map(lambda x: f.write(f"av{x}\n"), videos))
        if ranktype == "旧稿":
            videos = [x["AV号"] for x in xlsx_data]
            list(map(lambda x: f.write(f"{x.lower()}\n"), videos))
        if ranktype == "经典":
            videos = [x["AV号"] for x in xlsx_data]
            list(map(lambda x: f.write(f"{x.lower()}\n"), videos))
        if ranktype == "连续":
            videos = [x["aid"] for x in xlsx_data]
            list(map(lambda x: f.write(f"av{x}\n"), videos))
    if ranktype == "主榜":
        point_data = {
            x["排名"]: x["总分"]
            for x in xlsx_data[:21]
            if isinstance(x["排名"], int)  # and x["排名"] <= 21
        }
        # num_3 = len([x["排名"] for x in xlsx_data[:20] if x["排名"] <= 3])
        # num_10 = len([x["排名"] for x in xlsx_data[:20] if x["排名"] <= 10])
        json_data = [
            {
                "rank": n + 1,
                "true_rank": f"第{an2cn(x['排名'])}名" if n + 1 <= 3 else x["排名"],
                "video": f"./主榜视频/av{x['aid']}.mp4",
                # "text": f"./主榜{xlsx_data[2]['排名']}-1/Rank_{n+1}.png"
                "text": f"./主榜3-1/Rank_{n+1}.png" if n + 1 <= 3
                # else f"./主榜{xlsx_data[9]['排名']}-{xlsx_data[3]['排名']}/Rank_{n+1-3}.png"
                else f"./主榜10-4/Rank_{n+1-3}.png" if 3 < n + 1 <= 10
                # else f"./主榜{xlsx_data[19]['排名']}-{xlsx_data[10]['排名']}/Rank_{n+1-10}.png",
                else f"./主榜20-11/Rank_{n+1-10}.png",
                "delta": "+"
                + format(int(point_data[x["排名"]]) - int(point_data[x["排名"] + 1]), ",")
                if point_data.get(x["排名"] + 1) and n + 1 <= 3
                else "+000,000,000记得改数字！！"
                if not point_data.get(x["排名"] + 1) and n + 1 <= 3
                else "",
                "offset": 0,
            }
            for n, x in enumerate(xlsx_data[:20])
            # if x["排名"] <= 20
        ]
        data = dumps(json_data, indent=4, ensure_ascii=False)
        with open("data.json", "w", encoding="utf-8") as f:
            f.write(data)
    if ranktype == "旧稿":
        json_data = [
            {
                "rank": n + 1,
                "video": f"./主榜视频/{x['AV号'].lower()}.mp4",
                "text": f"./旧稿推荐/{n + 1}.png",
                "offset": 0,
            }
            for n, x in enumerate(xlsx_data)
        ]
        data = dumps(json_data, indent=4, ensure_ascii=False)
        with open("data_旧稿.json", "w", encoding="utf-8") as f:
            f.write(data)
    if ranktype == "经典":
        json_data = [
            {
                "rank": n + 1,
                "video": f"./主榜视频/{x['AV号'].lower()}.mp4",
                "text": f"./冷门推荐/{n + 1}.png",
                "offset": 0,
            }
            for n, x in enumerate(xlsx_data)
        ]
        json_data[-1]["text"] = "./经典推荐/1.png"
        json_data[-1]["rank"] = 0
        data = dumps(json_data, indent=4, ensure_ascii=False)
        with open("data_经典&冷门.json", "w", encoding="utf-8") as f:
            f.write(data)
    if ranktype == "连续":
        json_data = [
            {
                "rank": n + 1,
                "video": f"./主榜视频/av{x['aid']}.mp4",
                "text": f"./{weeks:03d}期连续在榜/Rank_{n + 1}.png",
                "offset": 0,
            }
            for n, x in enumerate(xlsx_data)
        ]
        data = dumps(json_data, indent=4, ensure_ascii=False)
        with open("data_连续.json", "w", encoding="utf-8") as f:
            f.write(data)


def main():
    print(f"\n\t现在是 {NOW.format('YYYY-MM-DD HH:MM:SS')}，本周应该是周刊第{an2cn(weeks)}期")
    print(f"\n\t将会查找文件名包含“主榜”“{weeks}”的Excel文件")
    print(f"\n\t将会查找文件名包含“连续在榜”“{weeks}”的Excel文件")
    print(f"\n\t将会查找文件名包含“旧稿回顾”“{an2cn(weeks)}”的Excel文件")
    print(f"\n\t将会查找文件名包含“经典回顾”“{an2cn(weeks)}”的Excel文件")

    input("\n\t回车继续执行...")

    with open("./psdownload/download.txt", "w", encoding="utf-8") as f:
        f.write("")

    main_excel = find(f"{weeks:03d}", "主榜")
    if main_excel is not None:
        xlsx2json(main_excel, "主榜")
        print("\tAE脚本数据“data.json”已经生成")

    continuity_excel = find(f"{weeks:03d}", "连续在榜")
    if continuity_excel is not None:
        xlsx2json(continuity_excel, "连续")
        print("\tAE脚本数据“data_连续.json”已经生成")

    old_excel = find(an2cn(weeks), "旧稿回顾")
    if old_excel is not None:
        xlsx2json(old_excel, "旧稿")
        print("\tAE脚本数据“data_旧稿.json”已经生成")

    classic_excel = find(an2cn(weeks), "经典回顾")
    if classic_excel is not None:
        xlsx2json(classic_excel, "经典")
        print("\tAE脚本数据“data_经典&冷门.json”已经生成")

    print(f"\n\t视频下载列表已保存至“{abspath('./psdownload/download.txt')}”")
    input("\t现在可以关闭本程序")


if __name__ == "__main__":
    main()
